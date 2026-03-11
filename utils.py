import streamlit as st
import pandas as pd
import io
import json
import os
import smtplib
import zipfile
import shutil
import uuid  # <-- Adicionado para gerar tokens únicos
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone


# --- CONSTANTES ---
ARQUIVO_PREFS_CLIENTES = "backup_prefs_clientes.json"
PLANILHA_PADRAO_ADMIN = "Levantamento_Base.xlsx"
PASTA_FOTOS = "fotos_uploads"

# Garantir que a pasta de fotos existe
if not os.path.exists(PASTA_FOTOS):
    os.makedirs(PASTA_FOTOS)


# --- PERSISTÊNCIA PREFERÊNCIAS DO CLIENTE ---

def carregar_prefs_todos_clientes():
    if os.path.exists(ARQUIVO_PREFS_CLIENTES):
        try:
            with open(ARQUIVO_PREFS_CLIENTES, "r") as f:
                return json.load(f)
        except:
            return {}
    return {}

def salvar_prefs_cliente(usuario, email, ativo):
    prefs = carregar_prefs_todos_clientes()
    ultimo_envio = None
    if usuario in prefs:
        ultimo_envio = prefs[usuario].get("ultimo_envio")

    prefs[usuario] = {
        "email": email,
        "ativo": ativo,
        "ultimo_envio": ultimo_envio
    }
    with open(ARQUIVO_PREFS_CLIENTES, "w") as f:
        json.dump(prefs, f, indent=4)

def atualizar_status_envio_cliente(usuario, data_envio):
    prefs = carregar_prefs_todos_clientes()
    if usuario in prefs:
        prefs[usuario]["ultimo_envio"] = data_envio
        with open(ARQUIVO_PREFS_CLIENTES, "w") as f:
            json.dump(prefs, f, indent=4)


# --- GERAÇÃO DE ZIP ---

def gerar_zip_usuario(usuario):
    try:
        nome_zip = f"backup_{usuario}_{datetime.now().strftime('%Y%m%d')}.zip"
        with zipfile.ZipFile(nome_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
            arquivo_dados = f"{usuario}.json" 
            if os.path.exists(arquivo_dados):
                zipf.write(arquivo_dados)
            
            pasta_fotos = os.path.join("fotos", usuario)
            if os.path.exists(pasta_fotos):
                for root, _, files in os.walk(pasta_fotos):
                    for file in files:
                        zipf.write(os.path.join(root, file), 
                                   arcname=os.path.join(usuario, "fotos", file))
        return nome_zip
    except Exception as e:
        print(f"[Erro ZIP Usuario] Falha ao gerar ZIP para {usuario}: {e}")
        return None

def gerar_zip_exportacao(dados_lista):
    if 'planilha_modelo' not in st.session_state: return None
    
    st.session_state['planilha_modelo'].seek(0)
    book = load_workbook(st.session_state['planilha_modelo'])
    
    for registro in dados_lista:
        if registro['tipo_equipamento'] in book.sheetnames:
            sheet = book[registro['tipo_equipamento']]
            headers = {sheet.cell(row=1, column=i).value: i for i in range(1, sheet.max_column + 1)}
            nova_linha_dados = registro['dados']
            next_row = sheet.max_row + 1
            
            for k, v in nova_linha_dados.items():
                if k in headers:
                    sheet.cell(row=next_row, column=headers[k], value=v)
            
            if "Fotos" in headers and "fotos" in registro:
                 nomes_fotos = ", ".join([f['nome_exportacao'] for f in registro["fotos"]])
                 sheet.cell(row=next_row, column=headers["Fotos"], value=nomes_fotos)
    
    excel_buffer = io.BytesIO()
    book.save(excel_buffer)
    excel_buffer.seek(0)
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("Levantamento_Cargas.xlsx", excel_buffer.getvalue())
        for registro in dados_lista:
            if "fotos" in registro and registro["fotos"]:
                uc = registro.get("cod_instalacao", "SemUC")
                tipo = registro.get("tipo_equipamento", "Geral")
                folder_path = f"Fotos/{uc} - {tipo}/"
                for foto in registro["fotos"]:
                    caminho_origem = foto["caminho_fisico"]
                    if os.path.exists(caminho_origem):
                        zf.write(caminho_origem, arcname=f"{folder_path}{foto['nome_exportacao']}")
    zip_buffer.seek(0)
    return zip_buffer


# --- DATAS E CAMINHOS ---

def get_data_hora_br():
    fuso_br = timezone(timedelta(hours=-3))
    return datetime.now(fuso_br)

def get_user_data_path(nome_usuario=None):
    user = nome_usuario if nome_usuario else st.session_state.get('usuario_ativo')
    if user:
        nome_limpo = "".join(filter(str.isalnum, user))
        return f"dados_{nome_limpo}.json"
    return None

def get_user_template_path(nome_usuario=None):
    user = nome_usuario if nome_usuario else st.session_state.get('usuario_ativo')
    if user:
        nome_limpo = "".join(filter(str.isalnum, user))
        return f"template_{nome_limpo}.xlsx"
    return None


# --- PERSISTÊNCIA JSON ---

def salvar_dados_locais(dados):
    path = get_user_data_path()
    if path:
        with open(path, "w") as f: json.dump(dados, f)

def carregar_dados_locais(path_especifico=None):
    path = path_especifico if path_especifico else get_user_data_path()
    if path and os.path.exists(path):
        with open(path, "r") as f: return json.load(f)
    return []


# --- LÓGICA DE FOTOS ---

def salvar_fotos_local(lista_fotos_obj, cod_instalacao):
    caminhos_salvos = []
    if not os.path.exists(PASTA_FOTOS):
        os.makedirs(PASTA_FOTOS)
    
    for item in lista_fotos_obj:
        arquivo = item['arquivo']
        nome_personalizado = item['nome']
        
        if arquivo:
            nome_orig = getattr(arquivo, "name", "foto_camera.jpg")
            ext = os.path.splitext(nome_orig)[1]
            if not ext: ext = ".jpg"
            
            nome_limpo = "".join(x for x in nome_personalizado if x.isalnum() or x in " -_")
            if not nome_limpo: nome_limpo = "imagem_sem_nome"
            
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S_%f")
            nome_arquivo_final = f"{cod_instalacao}_{timestamp}_{nome_limpo}{ext}"
            caminho_completo = os.path.join(PASTA_FOTOS, nome_arquivo_final)
            
            with open(caminho_completo, "wb") as f:
                f.write(arquivo.getbuffer())
            
            caminhos_salvos.append({
                "caminho_fisico": caminho_completo,
                "nome_exportacao": f"{nome_limpo}{ext}", 
                "nome_original": nome_orig
            })
            
    return caminhos_salvos


# --- FUNÇÕES AUXILIARES DE FORMATAÇÃO E ANÁLISE ---

def is_numeric_format(fmt_code):
    if not fmt_code: return False
    if str(fmt_code).lower() == 'general' or fmt_code == '@': return False
    return any(c in str(fmt_code) for c in ['0', '#', '%', 'E+'])

def is_list_numeric(lista_opcoes):
    if not lista_opcoes: return False
    try:
        [float(x) for x in lista_opcoes if x and str(x).strip() != '']
        return True
    except ValueError:
        return False

def analisar_modelo_excel(file_content):
    try:
        buffer = io.BytesIO(file_content) if isinstance(file_content, bytes) else io.BytesIO(file_content.getvalue())
        wb = load_workbook(buffer, data_only=True)
        estrutura = {}
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = []
            
            for col_idx in range(1, sheet.max_column + 1):
                c1 = sheet.cell(row=1, column=col_idx)
                c2 = sheet.cell(row=2, column=col_idx)
                
                nome_campo = None
                if c2.value: nome_campo = str(c2.value)
                elif c1.value: nome_campo = str(c1.value)
                
                if nome_campo:
                    tipo_detectado = "texto"
                    opcoes = []
                    sample_cell = sheet.cell(row=3, column=col_idx)
                    fmt_code = sample_cell.number_format
                    
                    has_validation = False
                    for dv in sheet.data_validations.dataValidation:
                        if sample_cell.coordinate in dv.sqref:
                            if dv.type == "list":
                                has_validation = True
                                formula = dv.formula1
                                if formula:
                                    opcoes_limpas = [op.strip() for op in formula.replace('"', '').split(',')]
                                    opcoes = opcoes_limpas
                                    is_numeric_list = is_list_numeric(opcoes_limpas)
                                    is_numeric_fmt = is_numeric_format(fmt_code)
                                    
                                    if is_numeric_list and is_numeric_fmt: tipo_detectado = "slider"
                                    elif dv.showErrorMessage is False: tipo_detectado = "selecao_aberta"
                                    else: tipo_detectado = "selecao"
                            break 

                    if not has_validation:
                        if is_numeric_format(fmt_code): tipo_detectado = "numero"
                        else: tipo_detectado = "texto"

                    headers.append({"nome": nome_campo, "col_letter": c1.column_letter, "tipo": tipo_detectado, "opcoes": opcoes})
            
            estrutura[sheet_name] = [{"nome": h["nome"], "tipo": h["tipo"], "opcoes": h["opcoes"]} for h in headers]
            
        return estrutura
    except Exception as e:
        st.error(f"Erro ao analisar o Excel: {e}")
        return {}
    
def carregar_modelo_atual():
    path_pessoal = get_user_template_path()
    if path_pessoal and os.path.exists(path_pessoal):
        with open(path_pessoal, "rb") as f:
            content = f.read()
            st.session_state['planilha_modelo'] = io.BytesIO(content)
            st.session_state['estrutura_modelo'] = analisar_modelo_excel(content)
            st.session_state['origem_modelo'] = "Pessoal"
    elif os.path.exists(PLANILHA_PADRAO_ADMIN):
        with open(PLANILHA_PADRAO_ADMIN, "rb") as f:
            content = f.read()
            st.session_state['planilha_modelo'] = io.BytesIO(content)
            st.session_state['estrutura_modelo'] = analisar_modelo_excel(content)
            st.session_state['origem_modelo'] = "Padrão do Sistema"


# --- EMAIL ---

def enviar_email(arquivo_buffer, destinatario, is_zip=False):
    try:
        msg = MIMEMultipart()
        msg['From'] = "levantamento.poupenergia@gmail.com"
        msg['To'] = destinatario
        msg['Subject'] = f"Levantamento {st.session_state['usuario_ativo']} - {get_data_hora_br().strftime('%d/%m/%Y')}"
        msg.attach(MIMEText("Segue em anexo o levantamento realizado.", 'plain'))
        
        part = MIMEBase('application', 'zip' if is_zip else 'octet-stream')
        part.set_payload(arquivo_buffer.getvalue())
        encoders.encode_base64(part)
        
        filename = "levantamento_completo.zip" if is_zip else "levantamento.xlsx"
        part.add_header('Content-Disposition', f'attachment; filename={filename}')
        
        msg.attach(part)
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login("levantamento.poupenergia@gmail.com", "kiqplowxqprcugjc")
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(e)
        return False

def enviar_email_backup_servico(destinatario, zip_path):
    try:
        msg = MIMEMultipart()
        msg['Subject'] = f"Backup Automático Diário - PoupEnergia - {datetime.now().strftime('%d/%m/%Y')}"
        msg['From'] = "levantamento.poupenergia@gmail.com"
        msg['To'] = destinatario
        msg.attach(MIMEText("Segue em anexo o backup completo do sistema.", 'plain'))
        
        with open(zip_path, "rb") as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(zip_path)}')
            msg.attach(part)
            
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login("levantamento.poupenergia@gmail.com", "kiqplowxqprcugjc")
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"Erro SMTP Backup: {e}")
        return False


def exportar_para_excel(dados):
    if not dados:
        return None
    try:
        lista_plana = []
        for item in dados:
            row = {
                "UC": item.get("cod_instalacao"),
                "Tipo": item.get("tipo_equipamento"),
                "Data": item.get("data_hora"),
            }
            dados_tecnicos = item.get("dados", {})
            if isinstance(dados_tecnicos, dict):
                row.update(dados_tecnicos)
            
            fotos = item.get("fotos", [])
            if fotos and isinstance(fotos, list):
                nomes_fotos = [f.get("nome_exportacao", "foto") for f in fotos]
                row["Fotos Anexadas"] = ", ".join(nomes_fotos)
            
            lista_plana.append(row)
            
        df = pd.DataFrame(lista_plana)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Exportacao")
            
        output.seek(0)
        return output

    except Exception as e:
        print(f"Erro ao exportar excel: {e}")
        return None


# --- CONFIGURAÇÕES DE BACKUP ---

def carregar_config_backup():
    if os.path.exists("config_backup.json"):
        with open("config_backup.json", "r") as f: return json.load(f)
    return {}

def salvar_config_backup(email):
    with open("config_backup.json", "w") as f: json.dump({"email": email}, f)

def carregar_estado_backup():
    if os.path.exists("estado_backup.json"):
        with open("estado_backup.json", "r") as f: return json.load(f)
    return {}

def salvar_estado_backup(data_envio, status):
    estado_atual = carregar_estado_backup()
    dados = {
        "data_ultimo_envio": data_envio if data_envio else estado_atual.get("data_ultimo_envio"),
        "status": status,
        "ultimo_check": datetime.now().isoformat()
    }
    with open("estado_backup.json", "w") as f: json.dump(dados, f)

def gerar_zip_sistema_completo():
    try:
        nome_arquivo = "sistema_poupenergia_backup.zip"
        arquivos = ["Levantamento_Base.xlsx", "db_formularios.json", "usuarios.json"]
        
        with zipfile.ZipFile(nome_arquivo, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for arq in arquivos:
                if os.path.exists(arq):
                    zipf.write(arq)
            if os.path.exists("fotos"):
                for root, _, files in os.walk("fotos"):
                    for file in files:
                        zipf.write(os.path.join(root, file))
        return nome_arquivo
    except:
        return None


# =========================================================================
# --- NOVO: CONTROLE DE SESSÃO PERSISTENTE SEGURA (VIA COOKIES) ---
# =========================================================================
ARQUIVO_SESSOES = "sessoes_ativas.json"
DIAS_EXPIRACAO = 7

def carregar_sessoes_ativas():
    """Carrega o banco de tokens do servidor."""
    if os.path.exists(ARQUIVO_SESSOES):
        try:
            with open(ARQUIVO_SESSOES, "r") as f:
                return json.load(f)
        except:
            return {}
    return {}

def salvar_sessoes_ativas(sessoes):
    with open(ARQUIVO_SESSOES, "w") as f:
        json.dump(sessoes, f)

def criar_sessao_persistente(usuario):
    """
    Cria um token aleatório e único para o dispositivo do usuário,
    salvando a referência no servidor e retornando o token.
    """
    sessoes = carregar_sessoes_ativas()
    token = str(uuid.uuid4()) # Gera um ID aleatório seguro
    expiracao = datetime.now() + timedelta(days=DIAS_EXPIRACAO)
    
    sessoes[token] = {
        "usuario": usuario,
        "expira_em": expiracao.strftime("%Y-%m-%d %H:%M:%S")
    }
    salvar_sessoes_ativas(sessoes)
    return token

def validar_token_sessao(token):
    """
    Recebe o token do cookie do navegador e verifica se ele
    ainda é válido no banco de sessões do servidor.
    """
    if not token:
        return None
        
    sessoes = carregar_sessoes_ativas()
    
    if token in sessoes:
        dados = sessoes[token]
        try:
            expira_em = datetime.strptime(dados["expira_em"], "%Y-%m-%d %H:%M:%S")
            if datetime.now() < expira_em:
                return dados["usuario"]
            else:
                # Token expirou, remove do servidor
                del sessoes[token]
                salvar_sessoes_ativas(sessoes)
                return None
        except:
            pass
            
    return None

def remover_sessao(token):
    """Invalida o token no servidor (Logout)."""
    if not token:
        return
    sessoes = carregar_sessoes_ativas()
    if token in sessoes:
        del sessoes[token]
        salvar_sessoes_ativas(sessoes)
